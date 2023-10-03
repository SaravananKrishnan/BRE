import openpyxl
from extractor import extract_business_rules
file_path = 'finaleval.xlsx'


workbook = openpyxl.load_workbook(file_path)

worksheet = workbook['newTable_1']

column_B_values = []
column_C_values = []
column_D_values = []

# till 22
for row_number in range(2, 22):
    print(f"\n\n\n\n\n\n\n-----------------------   row {row_number} ----------------------------------\n\n\n\n\n\n\n")
    value_B = worksheet.cell(row=row_number, column=2).value
    value_C = worksheet.cell(row=row_number, column=3).value
    value_D = worksheet.cell(row=row_number, column=4).value
    
    fp = './tests/{}'.format(value_B)
    prm = []
    sec = []
    if value_C == None:
        prm = []
    else:
        prm = [val.strip() for val in f"{value_C}".split('\n')]
    if value_D == None:
        prm = []
    else:
        sec = [val.strip() for val in f"{value_D}".split('\n')]
    with open('./input.txt', 'w') as ip:
        ip.write(str(len(prm))+'\n')
        for p in prm:
            ip.write(p + '\n')
        ip.write(str(len(sec))+'\n')
        for p in sec:
            ip.write(p + '\n')
    ans = extract_business_rules(fp)
    worksheet.cell(row=row_number,column=7,value=int(ans[8]))
    worksheet.cell(row=row_number,column=9,value=int(ans[2]))

workbook.save(file_path)

workbook.close()
