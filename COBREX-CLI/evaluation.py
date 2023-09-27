import openpyxl
from extractor import extract_business_rules
file_path = 'finaleval.xlsx'


workbook = openpyxl.load_workbook(file_path)

worksheet = workbook['Sheet1']

column_B_values = []
column_D_values = []
column_E_values = []

# till 22
for row_number in range(2, 22):
    print(f"\n\n\n\n\n\n\n-----------------------   row {row_number} ----------------------------------\n\n\n\n\n\n\n")
    value_B = worksheet.cell(row=row_number, column=2).value
    value_D = worksheet.cell(row=row_number, column=4).value
    value_E = worksheet.cell(row=row_number, column=5).value
    
    fp = './tests/{}'.format(value_B)
    prm = []
    sec = []
    if value_D == None:
        prm = []
    else:
        prm = [val.strip() for val in f"{value_D}".split('\n')]
    if value_E == None:
        prm = []
    else:
        sec = [val.strip() for val in f"{value_E}".split('\n')]
    with open('./input.txt', 'w') as ip:
        ip.write(str(len(prm))+'\n')
        for p in prm:
            ip.write(p + '\n')
        ip.write(str(len(sec))+'\n')
        for p in sec:
            ip.write(p + '\n')
    ans = extract_business_rules(fp)
    worksheet.cell(row=row_number,column=3,value=int(ans[0]))
    worksheet.cell(row=row_number,column=8,value=int(ans[1]))
    worksheet.cell(row=row_number,column=11,value=int(ans[2]))
    worksheet.cell(row=row_number,column=13,value=str(ans[3]))
    worksheet.cell(row=row_number,column=14,value=str(ans[4]))
    worksheet.cell(row=row_number,column=15,value=int(ans[5]))
    worksheet.cell(row=row_number,column=16,value=int(ans[6]))
    worksheet.cell(row=row_number,column=17,value=str(ans[7]))




workbook.save(file_path)

workbook.close()
